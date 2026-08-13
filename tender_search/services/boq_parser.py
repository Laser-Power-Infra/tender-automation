import os
import re
import json
import zipfile
import logging
import tempfile
import fnmatch

import openpyxl
import xlrd
import requests
from django.conf import settings
from tender_search.models import TenderMerged

logger = logging.getLogger(__name__)

DRIVE_FILE_ID_RE = re.compile(r"/file/d/([^/]+)/")


def extract_drive_file_id(url: str) -> str | None:
    m = DRIVE_FILE_ID_RE.search(url)
    return m.group(1) if m else None


def download_from_drive(file_id: str, dest_path: str) -> str:
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    session = requests.Session()

    response = session.get(url, stream=True)
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "")
    if "text/html" in content_type:
        first_chunk = response.iter_content(chunk_size=32768).__next__()
        text = first_chunk.decode("utf-8", errors="replace")
        m = re.search(r"confirm=([0-9A-Za-z\-_]+)", text)
        if m:
            confirm_token = m.group(1)
            url = f"https://drive.google.com/uc?export=download&confirm={confirm_token}&id={file_id}"
            response = session.get(url, stream=True)
            response.raise_for_status()

    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with open(dest_path, "wb") as f:
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)

    return dest_path


def extract_zip(zip_path: str, extract_to: str) -> list[str]:
    extracted_files = []

    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_to)

    for root, _dirs, files in os.walk(extract_to):
        for fname in files:
            fpath = os.path.join(root, fname)
            extracted_files.append(fpath)

            if fname.lower().endswith(".zip"):
                nested_dir = os.path.join(extract_to, fname.replace(".zip", "_extracted"))
                os.makedirs(nested_dir, exist_ok=True)
                nested_files = extract_zip(fpath, nested_dir)
                extracted_files.extend(nested_files)

    return extracted_files


def find_boq_file(file_paths: list[str]) -> str | None:
    for fp in file_paths:
        fname = os.path.basename(fp)
        name_no_ext = os.path.splitext(fname)[0]
        if fnmatch.fnmatch(name_no_ext.upper(), "*BOQ*"):
            return fp
    return None


def _normalize_header(val: str) -> str:
    return re.sub(r"[^a-z]", "", val.strip().lower())


_HEADER_KEYWORDS = {
    "description": ["description", "item", "particulars", "work", "nameofwork", "name of work", "material", "specification"],
    "quantity": ["qty", "quantity", "totalqty", "total qty", "estimated  rate"],
    "unit": ["unit", "uom", "measure"],
}


def _find_columns(cells_with_col):
    desc_col = None
    qty_col = None
    unit_col = None
    for val, col in cells_with_col:
        nv = _normalize_header(val)
        if desc_col is None:
            if nv in ("description", "itemdescription", "item"):
                desc_col = col
        if qty_col is None:
            if nv in ("quantity", "qty"):
                qty_col = col
        if unit_col is None:
            if nv in ("unit", "units", "uom"):
                unit_col = col
    return desc_col, qty_col, unit_col


def _parse_boq_xlsx(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found")

    header_row = None
    desc_col = qty_col = unit_col = None

    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        cells = [(str(c.value) if c.value is not None else "", c.column) for c in row]
        d, q, u = _find_columns(cells)
        if d is not None and (q is not None or u is not None):
            desc_col = d
            qty_col = q
            unit_col = u
            header_row = row[0].row
            break

    if header_row is None or desc_col is None:
        wb.close()
        raise ValueError("Could not find header row with description + quantity/unit columns")

    items = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        desc_val = row[desc_col - 1].value if desc_col <= len(row) else None
        desc = str(desc_val).strip() if desc_val is not None else ""
        if not desc or desc == "None" or _is_numeric(desc):
            continue

        qty_raw = row[qty_col - 1].value if qty_col and qty_col <= len(row) else None
        unit_raw = row[unit_col - 1].value if unit_col and unit_col <= len(row) else None

        qty = _to_number(qty_raw)
        items.append({
            "description": desc,
            "quantity": qty if qty is not None else "no_quantity_available",
            "unit": str(unit_raw).strip() if unit_raw is not None else "",
        })

    wb.close()

    if not items:
        raise ValueError("No data rows found under header")

    return items


def _parse_boq_xls(file_path: str) -> list[dict]:
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)

    header_row = None
    desc_col = qty_col = unit_col = None

    for r in range(min(20, ws.nrows)):
        cells = [(str(ws.cell_value(r, c)), c) for c in range(ws.ncols)]
        d, q, u = _find_columns(cells)
        if d is not None and (q is not None or u is not None):
            desc_col = d
            qty_col = q
            unit_col = u
            header_row = r
            break

    if header_row is None or desc_col is None:
        raise ValueError("Could not find header row with description + quantity/unit columns")

    items = []
    for r in range(header_row + 1, ws.nrows):
        desc = str(ws.cell_value(r, desc_col)).strip()
        if not desc or desc == "None" or _is_numeric(desc):
            continue

        qty_raw = ws.cell_value(r, qty_col) if qty_col is not None else None
        unit_raw = ws.cell_value(r, unit_col) if unit_col is not None else None

        qty = _to_number(qty_raw)
        items.append({
            "description": desc,
            "quantity": qty if qty is not None else "no_quantity_available",
            "unit": str(unit_raw).strip() if unit_raw is not None else "",
        })

    if not items:
        raise ValueError("No data rows found under header")

    return items


def parse_boq(file_path: str) -> list[dict]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        return _parse_boq_xls(file_path)
    elif ext == ".xlsx":
        return _parse_boq_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext} (expected .xls or .xlsx)")


def _to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val == val else None
    s = str(val).strip().replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _is_numeric(val: str) -> bool:
    try:
        float(val.strip())
        return True
    except ValueError:
        return False


def save_boq_to_db(reference_no: str, items: list[str]) -> dict:
    tender = TenderMerged.objects.filter(referenceno=reference_no).first()
    if not tender:
        return {"success": False, "error": f"TenderMerged not found for reference_no={reference_no}"}

    tender.size = json.dumps(items)
    tender.parsestatus = "COMPLETED"
    tender.parseerror = None
    tender.save()
    logger.info("Saved BOQ items for %s (size field)", reference_no)
    print(json.dumps(items))
    return {"success": True}


def _save_failure_to_db(reference_no: str, error_msg: str):
    try:
        tender = TenderMerged.objects.filter(referenceno=reference_no).first()
        if tender:
            tender.parsestatus = "FAILED"
            tender.parseerror = error_msg
            tender.save()
            logger.info("Updated parsestatus=FAILED for %s", reference_no)
    except Exception as e:
        logger.warning("Could not save failure to DB for %s: %s", reference_no, e)


def process_boq(reference_no: str, drive_link: str) -> dict:
    result = {
        "reference_no": reference_no,
        "success": False,
        "boq_file": None,
        "items": [],
        "error": None,
    }

    file_id = extract_drive_file_id(drive_link)
    if not file_id:
        result["error"] = f"Could not extract file ID from Drive link: {drive_link}"
        return result

    temp_dir = getattr(settings, "TENDER_PARSING_TEMP_DIR", tempfile.gettempdir())
    safe_name = reference_no.replace("/", "-").replace("\\", "-")
    zip_path = os.path.join(temp_dir, f"{safe_name}.zip")
    extract_dir = os.path.join(temp_dir, f"{safe_name}_extracted")

    try:
        logger.info("Downloading file_id=%s -> %s", file_id, zip_path)
        download_from_drive(file_id, zip_path)

        logger.info("Extracting %s -> %s", zip_path, extract_dir)
        os.makedirs(extract_dir, exist_ok=True)
        extracted_files = extract_zip(zip_path, extract_dir)

        boq_path = find_boq_file(extracted_files)
        if not boq_path:
            result["error"] = "No BOQ file found among extracted files"
            _save_failure_to_db(reference_no, result["error"])
            return result

        result["boq_file"] = os.path.basename(boq_path)
        logger.info("Found BOQ file: %s", boq_path)

        items = parse_boq(boq_path)
        result["items"] = [
            f"{i}. {item['description']}_{item['quantity']}_{item['unit']}"
            for i, item in enumerate(items, start=1)
        ]
        result["success"] = True

        db_result = save_boq_to_db(reference_no, result["items"])
        result["db_save"] = db_result

    except Exception as e:
        logger.exception("BOQ processing failed")
        result["error"] = str(e)
        _save_failure_to_db(reference_no, str(e))

    finally:
        if os.path.exists(zip_path):
            try:
                os.remove(zip_path)
            except Exception as e:
                logger.warning("Could not delete %s: %s", zip_path, e)
        if os.path.exists(extract_dir):
            try:
                import shutil
                shutil.rmtree(extract_dir, ignore_errors=True)
            except Exception as e:
                logger.warning("Could not delete %s: %s", extract_dir, e)

    return result
