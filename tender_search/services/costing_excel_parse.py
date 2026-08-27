import logging
import os,json
import re
import tempfile
from datetime import datetime

import openpyxl
import requests
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from pprint import pprint
from tender_search.models import Costingsheetdetails, Items, TenderMerged

logger = logging.getLogger(__name__)

# Drive URL patterns: /file/d/<id>/ and ?id=<id> (open?id=, uc?export=download&id=)
_DRIVE_FILE_ID_RE_1 = re.compile(r"/file/d/([^/?&#]+)")
_DRIVE_FILE_ID_RE_2 = re.compile(r"[?&]id=([^&#]+)")


def _extract_drive_file_id(url: str) -> str | None:
    m = _DRIVE_FILE_ID_RE_1.search(url)
    if m:
        return m.group(1)
    m = _DRIVE_FILE_ID_RE_2.search(url)
    return m.group(1) if m else None


def _is_drive_link(url: str) -> bool:
    return "drive.google.com" in url.lower()


def _format_http_error(err) -> str:
    """Return detailed HttpError string including status, reason, json body."""
    try:
        from googleapiclient.errors import HttpError
        import json as _json
        if isinstance(err, HttpError):
            status = getattr(getattr(err, "resp", None), "status", "")
            reason = getattr(getattr(err, "resp", None), "reason", "")
            body = ""
            try:
                raw = err.content.decode("utf-8", errors="replace") if getattr(err, "content", None) else ""
                # Try pretty json
                try:
                    parsed = _json.loads(raw)
                    body = _json.dumps(parsed)
                except Exception:
                    body = raw[:1000]
            except Exception:
                body = str(err)[:1000]
            return f"HttpError {status} {reason} - {body}"
    except Exception:
        pass
    return f"{type(err).__name__}: {err}"


def _get_service_account_service():
    """Build service with service-account only (bypass OAuth token) for fallback."""
    from google.oauth2.service_account import Credentials as ServiceAccountCredentials
    from googleapiclient.discovery import build

    sa_creds = ServiceAccountCredentials.from_service_account_info(
        {
            "client_email": settings.GDRIVE_CLIENT_EMAIL,
            "private_key": settings.GDRIVE_PRIVATE_KEY,
            "token_uri": "https://oauth2.googleapis.com/token",
        },
        scopes=[
            "https://www.googleapis.com/auth/drive.file",
            "https://www.googleapis.com/auth/drive.readonly",
        ],
    )
    return build("drive", "v3", credentials=sa_creds)


def _download_via_api(service, file_id: str, dest_path: str) -> None:
    from googleapiclient.http import MediaIoBaseDownload
    import io

    # Detect native Google Workspace files that need export
    try:
        meta = service.files().get(fileId=file_id, fields="mimeType,name").execute()
        mime = meta.get("mimeType", "")
    except Exception as e:
        logger.warning("Drive files().get failed for %s: %s", file_id, _format_http_error(e))
        mime = ""
    os.makedirs(os.path.dirname(dest_path) or ".", exist_ok=True)
    if mime == "application/vnd.google-apps.spreadsheet":
        request = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = service.files().get_media(fileId=file_id)
    fh = io.FileIO(dest_path, "wb")
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.close()
    # Verify not HTML
    with open(dest_path, "rb") as check:
        head = check.read(512)
        if head.lstrip().lower().startswith(b"<!doctype html") or b"accounts.google.com" in head.lower():
            preview = head[:300].decode("utf-8", errors="replace").replace("\n", " ")
            try:
                os.remove(dest_path)
            except OSError:
                pass
            raise ValueError(f"Expected Excel, got HTML after API download. Preview: {preview}")


def _download_drive_file(file_id: str, dest_path: str) -> None:
    """Download a Google Drive file. Tries authenticated Drive API first, falls back to uc endpoint. Surfaces real HttpError."""
    api_error_detail = None
    last_api_err = None

    # --- 1) Authenticated API path: try service-account (drive.readonly, can read files shared to it)
    #        FIRST, then OAuth/user token. The token.json OAuth credential has drive.file scope only,
    #        which returns 404 for files not created by the app (e.g. costing files owned by samrat.dey).
    services_to_try = []
    # Service-account first - proven to read files shared to it (drive.file + drive.readonly scopes)
    try:
        if getattr(settings, "GDRIVE_CLIENT_EMAIL", "") and getattr(settings, "GDRIVE_PRIVATE_KEY", ""):
            services_to_try.append(("service-account", _get_service_account_service()))
    except Exception as e:
        logger.warning("Failed to build service-account service for %s: %s", file_id, _format_http_error(e))
    # OAuth / user token fallback (drive.file scope - only works for files the app itself created)
    try:
        from tender_search.services.google_drive import _get_authenticated_service
        services_to_try.append(("oauth/service", _get_authenticated_service()))
    except Exception as e:
        logger.warning("Failed to get OAuth/service from _get_authenticated_service for %s: %s", file_id, _format_http_error(e))

    for label, service in services_to_try:
        try:
            _download_via_api(service, file_id, dest_path)
            logger.info("Downloaded Drive file %s via %s API -> %s", file_id, label, dest_path)
            return
        except Exception as api_err:
            err_str = str(api_err)
            if "Expected Excel, got HTML after API download" in err_str:
                raise
            api_error_detail = _format_http_error(api_err)
            last_api_err = api_err
            logger.warning("Authenticated Drive download failed for %s via %s: %s", file_id, label, api_error_detail)
            # Clean partial file if created
            if os.path.exists(dest_path):
                try:
                    if os.path.getsize(dest_path) < 512:
                        os.remove(dest_path)
                except OSError:
                    pass
            # If 404, try next credential before falling back
            continue

    # If all API attempts failed, keep api_error_detail for final message
    if api_error_detail:
        logger.warning("All authenticated attempts failed for %s: %s — falling back to uc endpoint", file_id, api_error_detail)
    else:
        logger.warning("No authenticated service available for %s — falling back to uc endpoint", file_id)

    # --- 2) Fallback: unauthenticated uc?export=download (for public anyone-with-link files) ---
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    session = requests.Session()
    resp = session.get(url, stream=True, timeout=120)
    resp.raise_for_status()

    content_type = resp.headers.get("Content-Type", "")
    # Large files return HTML warning page; check both header and body
    if "text/html" in content_type.lower():
        # Peek first chunk without consuming entire stream for confirm extraction
        first_chunk = b""
        for chunk in resp.iter_content(chunk_size=32768):
            if chunk:
                first_chunk = chunk
                break
        text = first_chunk.decode("utf-8", errors="replace") if first_chunk else ""
        m = re.search(r"confirm=([0-9A-Za-z-_]+)", text)
        if m:
            confirm_token = m.group(1)
            url = f"https://drive.google.com/uc?export=download&confirm={confirm_token}&id={file_id}"
            resp = session.get(url, stream=True, timeout=120)
            resp.raise_for_status()
        else:
            # Private / not shared file redirects to sign-in HTML
            if "accounts.google.com" in text or "<!doctype html" in text.lower():
                preview = text[:500].replace("\n", " ")
                # Include real API error detail instead of only hardcoded text
                api_part = f" API error: {api_error_detail}" if api_error_detail else ""
                sa_email = getattr(settings, 'GDRIVE_CLIENT_EMAIL', '')
                raise ValueError(
                    f"Drive file not accessible (private or not shared) for {file_id}.{api_part} "
                    f"uc fallback got sign-in HTML. Ensure file is shared with {sa_email} or made anyone-with-link. "
                    f"Preview: {preview}"
                ) from last_api_err
            # If HTML but no confirm token and not sign-in, treat as error
            if text.strip():
                preview = text[:500].replace("\n", " ")
                raise ValueError(f"Expected Excel, got HTML from Drive. Preview: {preview}")

    os.makedirs(os.path.dirname(dest_path) or ".", exist_ok=True)
    with open(dest_path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)

    # Post-write HTML sniff: Drive may return HTML with 200 + octet-stream mislabel
    try:
        with open(dest_path, "rb") as fh:
            head = fh.read(512)
            if head.lstrip().lower().startswith(b"<!doctype html") or b"accounts.google.com" in head.lower():
                preview = head[:200].decode("utf-8", errors="replace")
                raise ValueError(f"Expected Excel, got HTML after download. Preview: {preview}")
    except ValueError:
        # Remove invalid file
        try:
            os.remove(dest_path)
        except OSError:
            pass
        raise

def _to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val == val else None
    s = str(val).strip().replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _strip_pct(val):
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        if abs(val) < 10:
            val = val * 100
        return f"{val:.2f}"
    s = str(val).strip().replace("%", "").strip()
    return s if s else ""


def _cell(row_obj, col_idx):
    if col_idx is None:
        return None
    idx = col_idx - 1
    return row_obj[idx].value if idx < len(row_obj) else None


def _has_value(raw):
    if raw is None:
        return False
    if isinstance(raw, str):
        return raw.strip() != ""
    if isinstance(raw, (int, float)):
        return raw != 0
    return True


NON_MATERIAL_KEYWORDS = [
    "docket", "base date", "item code", "erp item", "propose erp",
    "ex-works", "freight", "price including", "gst", "qty", "quantity",
    "unit", "total price", "non chain bom", "drum type", "drum code",
    "drum size", "party standard", "mfg", "prof", "1st ex-works",
    "intt", "insp", "ttr charges",
]


def _is_non_material_header(header_lower):
    return any(kw in header_lower for kw in NON_MATERIAL_KEYWORDS)


def _find_all_header_rows(sheet):
    headers = []
    rcount = 0
    # print(sheet)
    for row in sheet.iter_rows():
        # print(row)
        vals = [str(c.value).strip() if c.value else "" for c in row]
        # print(f"count: {rcount+1} ----- {vals}")
        has_docket = any("docket" in v.lower() for v in vals)
        has_erp = any("propose" in v.lower() and "erp" in v.lower() for v in vals)
        if has_docket and has_erp:
            headers.append(row[0].row)
        rcount+=1
    return headers


def _build_column_map(sheet, header_row):
    col_map = {}
    for cell in sheet[header_row]:
        if cell.value is None:
            continue
        val = str(cell.value).strip().lower()
        if "docket" in val:
            col_map.setdefault("docket_no", cell.column)
        if "propose" in val and "erp" in val:
            col_map.setdefault("propose_erp", cell.column)
        if "qty" in val or "quantity" in val:
            col_map.setdefault("qty", cell.column)
        if "item code" in val:
            col_map.setdefault("item_code", cell.column)
        if "base date" in val:
            col_map.setdefault("base_date", cell.column)
        if "total price" in val:
            col_map.setdefault("total_price", cell.column)
    return col_map


def _find_cva_columns(ws, header_row_num):
    cols = {}
    for cell in ws[header_row_num]:
        if cell.value is None:
            continue
        v = str(cell.value).strip().lower()
        if "mfg" in v and "%" in v:
            cols.setdefault("mfg", cell.column)
        if "intt" in v and "%" in v:
            cols.setdefault("intt", cell.column)
        if "insp" in v and "%" in v:
            cols.setdefault("insp", cell.column)
        if "prof" in v and "%" in v:
            cols.setdefault("prof", cell.column)
        if "ttr" in v and "%" in v:
            cols.setdefault("ttr", cell.column)
    return cols


_FIRM_RE = re.compile(r"\bfirm\b", re.IGNORECASE)
_VARIABLE_RE = re.compile(r"\bvariable\b", re.IGNORECASE)
_IEEMA_RE = re.compile(r"\bieema\b", re.IGNORECASE)
_CACMAI_RE = re.compile(r"\bcacmai\b", re.IGNORECASE)


def _find_price_basis(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            v = str(cell.value)
            if _FIRM_RE.search(v):
                return "FIRM"
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            v = str(cell.value)
            if _VARIABLE_RE.search(v):
                return "VARIABLE"
    return "FIRM"


def _find_applicable_index(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            v = str(cell.value)
            if _IEEMA_RE.search(v):
                return "IEEMA"
            if _CACMAI_RE.search(v):
                return "CACMAI"
    return None


@transaction.atomic
def save_costing_to_db(gemid, table_data, price_basis="FIRM", applicable_index=None):
    merged = TenderMerged.objects.filter(referenceno=gemid).first()
    if not merged:
        return {"success": False, "error": f"TenderMerged not found for docketno={gemid}"}

    materials = {}
    for t in table_data:
        for k, v in t.get("materials", {}).items():
            if k not in materials:
                materials[k] = v

    merged.rawmaterials = json.dumps(materials)
    merged.price = price_basis
    if applicable_index:
        merged.applicableindex = applicable_index
    if table_data:
        bd = table_data[0].get("base_date")
        if bd:
            merged.basedate = datetime.combine(bd, datetime.min.time())
    our_value = sum(
        row.get("total_price") or 0
        for t in table_data
        for row in t.get("rows", [])
    )
    merged.ourvalue = str(our_value) if our_value else None
    merged.save()

    all_codes = {
        row.get("item_code")
        for t in table_data
        for row in t.get("rows", [])
        if row.get("item_code")
    }
    schedule_map = dict(
        Items.objects.filter(itemcode__in=all_codes).values_list("itemcode", "itemschedule")
    )

    now = timezone.now()
    for t in table_data:
        for row in t.get("rows", []):
            item_code = (row.get("item_code") or "").strip()
            if not item_code:
                continue
            location = row.get("location") or ""
            obj = Costingsheetdetails.objects.filter(
                tendermergedid=merged, itemcode=item_code, location=location
            ).first()
            if obj is None and not location:
                obj = Costingsheetdetails.objects.filter(
                    tendermergedid=merged, itemcode=item_code
                ).first()
            fields = {
                "proposederpitemname": row.get("item_name"),
                "proposederpquantity": (
                    str(row.get("quantity")) if row.get("quantity") is not None else None
                ),
                "priceoffullquantity": (
                    str(row.get("total_price")) if row.get("total_price") is not None else None
                ),
                "cva": row.get("cva"),
                "location": location,
                "itemschedule": schedule_map.get(item_code),
                "updatedat": now,
            }
            if obj:
                for k, v in fields.items():
                    setattr(obj, k, v)
                obj.save()
            else:
                Costingsheetdetails.objects.create(
                    itemcode=item_code,
                    tendermergedid=merged,
                    createdat=now,
                    **fields,
                )
    return {"success": True}


def _build_laser_cost_payload(table_data, price_basis):
    materials = {}
    for t in table_data:
        for k, v in t.get("materials", {}).items():
            if k not in materials:
                materials[k] = v

    docket_no = None
    costing_sheet_details = []
    for t in table_data:
        if docket_no is None:
            docket_no = t.get("docket_no")
        for row in t.get("rows", []):
            costing_sheet_details.append({
                "proposedErpItemName": row.get("item_name"),
                "proposedQty": row.get("quantity"),
                "cvaValue": row.get("cva"),
            })

    return {
        "docketNo": docket_no,
        "priceBasis": price_basis,
        "rawMaterials": materials,
        "costingSheetDetails": costing_sheet_details,
    }


def _send_to_laser_cost_api(payload):
    url = f"http://{settings.LASER_TENDER_COST_API}/api/costing/parsed"
    headers = {"Authorization": f"Bearer {settings.WORKER_API_KEY}"}
    resp = requests.post(url, json=payload, headers=headers, timeout=60)
    resp.raise_for_status()
    return resp.json()


def _parse_table(ws, header_row_num, end_row):
    col_map = _build_column_map(ws, header_row_num)
    erp_col = col_map.get("propose_erp")
    if erp_col is None:
        return None
    docket_col = col_map.get("docket_no")
    if docket_col is None:
        return None

    has_docket_data = False
    for row_idx in range(header_row_num + 1, end_row):
        row_obj = list(ws[row_idx])
        dv = _cell(row_obj, docket_col)
        if dv is not None and str(dv).strip() != "":
            has_docket_data = True
            break
    if not has_docket_data:
        return None

    cva_cols = _find_cva_columns(ws, header_row_num)
    if not cva_cols:
        return None

    header_cells = list(ws[header_row_num])
    rate_row_num = None
    material_rates = {}
    rate_columns = {}

    for row_idx in range(header_row_num + 1, end_row):
        row_obj = list(ws[row_idx])
        erp_val = _cell(row_obj, erp_col)
        has_erp = erp_val is not None and str(erp_val).strip() != ""
        if has_erp:
            continue
        material_count = 0
        non_material_count = 0
        for i, cell in enumerate(row_obj):
            if i >= len(header_cells):
                break
            hdr = str(header_cells[i].value).strip() if header_cells[i].value else ""
            num = _to_number(cell.value)
            if num is not None and abs(num) > 0:
                if _is_non_material_header(hdr.lower()):
                    non_material_count += 1
                else:
                    material_count += 1
        if material_count >= 1 and non_material_count == 0:
            rate_row_num = row_idx
            break

    if rate_row_num is not None:
        rate_row = list(ws[rate_row_num])
        for i, cell in enumerate(rate_row):
            if i >= len(header_cells):
                break
            hdr = str(header_cells[i].value).strip() if header_cells[i].value else ""
            if not hdr or _is_non_material_header(hdr.lower()):
                continue
            num = _to_number(cell.value)
            if num is not None and abs(num) > 0:
                material_rates[hdr] = num
                rate_columns[hdr] = cell.column

    rows = []
    table_base_date = None
    table_docket_no = None
    used_materials = set()

    location_col = None
    nxt_header = _cell(header_cells, docket_col + 1)
    if nxt_header is None or str(nxt_header).strip() == "":
        for row_idx in range(header_row_num + 1, end_row):
            rv = _cell(list(ws[row_idx]), docket_col + 1)
            if rv is not None and isinstance(rv, str) and rv.strip():
                location_col = docket_col + 1
                break
    item_code_col = col_map.get("item_code")
    base_date_col = col_map.get("base_date")
    total_price_col = col_map.get("total_price")

    for row_idx in range(header_row_num + 1, end_row):
        row_obj = list(ws[row_idx])
        docket_val = _cell(row_obj, docket_col)
        has_docket = docket_val is not None and str(docket_val).strip() != ""
        if not has_docket:
            continue
        if table_docket_no is None:
            table_docket_no = str(docket_val).strip()
        erp_val = _cell(row_obj, erp_col)
        has_erp = erp_val is not None and str(erp_val).strip() != ""
        if not has_erp:
            continue

        has_any_cva = any(
            _has_value(_cell(row_obj, cva_cols.get(k)))
            for k in ("mfg", "intt", "insp", "prof", "ttr")
            if cva_cols.get(k) is not None
        )
        if not has_any_cva:
            continue

        name = str(erp_val).strip()
        location_val = _cell(row_obj, location_col)
        location = str(location_val).strip() if location_val is not None else ""
        item_code_val = _cell(row_obj, item_code_col)
        item_code = str(item_code_val).strip() if item_code_val is not None else ""
        qty_val = _to_number(_cell(row_obj, col_map.get("qty")))

        if table_base_date is None:
            bd = _cell(row_obj, base_date_col)
            if bd is not None:
                if isinstance(bd, datetime):
                    table_base_date = bd.date()
                else:
                    s = str(bd).strip()
                    parsed = None
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                        try:
                            parsed = datetime.strptime(s, fmt).date()
                            break
                        except ValueError:
                            continue
                    table_base_date = parsed

        total_price_val = _to_number(_cell(row_obj, total_price_col))
        if total_price_val is not None:
            total_price_val = round(total_price_val, 2)

        if cva_cols:
            parts = []
            for k in ("mfg", "intt", "insp", "prof", "ttr"):
                col = cva_cols.get(k)
                if col is None:
                    continue
                raw = _cell(row_obj, col)
                if _has_value(raw):
                    parts.append(_strip_pct(raw))
            cva_str = "@".join(parts)
        else:
            cva_str = ""

        rows.append({
            "item_code": item_code,
            "item_name": name,
            "quantity": qty_val,
            "total_price": total_price_val,
            "location": location,
            "cva": cva_str,
        })

        for hdr, col in rate_columns.items():
            mval = _to_number(_cell(row_obj, col))
            if mval is not None and abs(mval) > 0:
                used_materials.add(hdr)

    if not rows:
        return None

    material_rates = {
        hdr: val for hdr, val in material_rates.items() if hdr in used_materials
    }

    header = [str(c.value).strip() if c.value else "" for c in header_cells]
    while header and header[-1] == "":
        header.pop()

    return {
        "base_date": table_base_date,
        "docket_no": table_docket_no,
        "rows": rows,
        "materials": material_rates,
        "header": header,
    }


def parse_costing_excel(gemid, appsheet_link, sender=None):
    temp_dir = tempfile.gettempdir()
    safe_id = str(gemid).replace("/", "-").replace("\\", "-")
    temp_path = os.path.join(temp_dir, f"costing_{safe_id}.xlsx")
    source_is_url = bool(re.match(r"^https?://", str(appsheet_link).strip().lower()))
    downloaded = False

    try:
        if source_is_url:
            if _is_drive_link(appsheet_link):
                file_id = _extract_drive_file_id(appsheet_link)
                if not file_id:
                    return {"gemid": gemid, "error": f"Could not extract Drive file ID from link: {appsheet_link}"}
                _download_drive_file(file_id, temp_path)
                downloaded = True
                source_path = temp_path
            else:
                resp = requests.get(appsheet_link, stream=True, timeout=120)
                resp.raise_for_status()

                content_type = resp.headers.get("Content-Type", "")
                if "html" in content_type.lower():
                    first_200 = resp.content[:200].decode("utf-8", errors="replace")
                    return {"gemid": gemid, "error": f"Expected Excel, got HTML. Preview: {first_200}"}

                # Sniff body even if content-type is octet-stream (Drive sometimes mislabels)
                head = resp.content[:512] if hasattr(resp, "content") else b""
                if head.lstrip().lower().startswith(b"<!doctype html") or b"accounts.google.com" in head.lower():
                    first_200 = head[:200].decode("utf-8", errors="replace")
                    return {"gemid": gemid, "error": f"Expected Excel, got HTML. Preview: {first_200}"}

                with open(temp_path, "wb") as f:
                    f.write(resp.content)
                downloaded = True
                source_path = temp_path
        else:
            source_path = appsheet_link
            if not os.path.exists(source_path):
                return {"gemid": gemid, "error": f"Costing file not found at path: {source_path}"}

        wb = openpyxl.load_workbook(source_path, data_only=True)

        calc_sheets = [ws for ws in wb.worksheets if "AUTO CALCULATION SHEET" in ws.title.upper()]
        if not calc_sheets:
            return {"gemid": gemid, "error": "AUTO CALCULATION SHEET tab not found"}

        tables = []
        for ws in calc_sheets:
            header_rows = _find_all_header_rows(ws)
            if not header_rows:
                continue
            for i, hr in enumerate(header_rows):
                end_row = header_rows[i + 1] if i + 1 < len(header_rows) else ws.max_row + 1
                table = _parse_table(ws, hr, end_row)
                if table:
                    tables.append(table)

        if not tables:
            return {"gemid": gemid, "error": "No valid table headers found (DOCKET NO + PROPOSE ERP)"}

        price_basis = _find_price_basis(calc_sheets[0])
        applicable_index = _find_applicable_index(calc_sheets[0])
        if sender == "laser_cost":
            payload = _build_laser_cost_payload(tables, price_basis)
            pprint(payload)
            _send_to_laser_cost_api(payload)
        else:
            save_costing_to_db(gemid=gemid, table_data=tables, price_basis=price_basis, applicable_index=applicable_index)
        result = {
            "gemid": gemid,
            "tables": tables,
            "price": price_basis,
            "applicableIndex": applicable_index,
            "sender": sender,
        }
        logger.info("Costing parse result for %s: %s", gemid, json.dumps(result, default=str))
        return result

    except Exception as e:
        return {"gemid>>>>": gemid, "error": str(e)}

    finally:
        if downloaded and os.path.exists(temp_path):
            os.remove(temp_path)
